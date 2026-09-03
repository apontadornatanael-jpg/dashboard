import streamlit as st
import sqlite3
import pandas as pd
import re
import hashlib
from datetime import date, datetime, time, timedelta
from io import BytesIO
from pathlib import Path
import tempfile
import os

# ============================================================
# CONFIGURAÇÃO
# ============================================================
BASE_DIR = Path(__file__).resolve().parent
DB = str(BASE_DIR / "ddh.db")
BACKUP_DIR = BASE_DIR / "backups"
BACKUP_DIR.mkdir(exist_ok=True)
LOGO_PATH = BASE_DIR / "logo_ddh.png"

st.set_page_config(
    page_title="DDH Campo",
    page_icon="⛏️",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>

/* =====================================================
   CABEÇALHO DO STREAMLIT - MANTER VISÍVEL
   ===================================================== */

header[data-testid="stHeader"] {
    display: block !important;
    visibility: visible !important;
    opacity: 1 !important;
    height: 3rem !important;
    background: transparent !important;
    z-index: 999999 !important;
}

/* Botão de abrir a barra lateral */
[data-testid="stSidebarCollapsedControl"] {
    display: flex !important;
    visibility: visible !important;
    opacity: 1 !important;
    position: fixed !important;
    top: 0.5rem !important;
    left: 0.5rem !important;
    z-index: 9999999 !important;
}

/* Remove apenas o menu principal */
#MainMenu {
    display: none !important;
}

/* Ajuste do conteúdo */
.block-container {
    padding-top: 1rem;
}

</style>
""", unsafe_allow_html=True)
# ============================================================
# SESSION STATE
# ============================================================
DEFAULT_SESSION = {
    "logado": False,
    "usuario_id": None,
    "usuario_nome": None,
    "usuario_login": None,
    "usuario_nivel": None,
    "page": "🏠 Painel DDH",
    "boletim_edit_id": None,
}
for k, v in DEFAULT_SESSION.items():
    if k not in st.session_state:
        st.session_state[k] = v

PAGES_ADMIN = [
    "🏠 Painel DDH",
    "📝 Novo Boletim",
    "📅 Produção Diária",
    "📊 Relatórios e Excel",
    "📋 Boletins Salvos",
    "👷 Colaboradores",
    "👥 Equipes",
    "🔩 Sondas",
    "⚙️ Cadastros",
    "🛠️ Gerenciamento",
    "👤 Usuários",
    "💾 Backup",
]

PAGES_SUPERVISOR = [
    "🏠 Painel DDH",
    "📝 Novo Boletim",
    "📅 Produção Diária",
    "📊 Relatórios e Excel",
    "📋 Boletins Salvos",
    "👷 Colaboradores",
    "👥 Equipes",
    "🔩 Sondas",
    "⚙️ Cadastros",
    "🛠️ Gerenciamento",
]

PAGES_CAMPO = [
    "📝 Novo Boletim",
    "📋 Boletins Salvos",
    "🕳️ Cadastro de Furos",
    "📖 Consultar Atividades",
]

# ============================================================
# FUNÇÕES BÁSICAS
# ============================================================
def hash_senha(senha):
    return hashlib.sha256(senha.encode("utf-8")).hexdigest()

def verificar_senha(senha, senha_hash):
    return hash_senha(senha) == senha_hash

def conn():
    c = sqlite3.connect(DB, check_same_thread=False, timeout=30)
    c.execute("PRAGMA foreign_keys = ON")
    return c

# ============================================================
# BACKUP DO BANCO SQLITE
# ============================================================
def criar_backup(prefixo="backup"):
    """Cria um backup consistente do SQLite usando a API nativa de backup."""
    if not Path(DB).exists():
        return None
    agora = datetime.now().strftime("%Y%m%d_%H%M%S")
    destino = BACKUP_DIR / f"{prefixo}_ddh_{agora}.db"
    origem = sqlite3.connect(DB, timeout=30)
    destino_conn = sqlite3.connect(str(destino))
    try:
        origem.backup(destino_conn)
        destino_conn.commit()
    finally:
        destino_conn.close()
        origem.close()

    # Mantém os 30 backups automáticos mais recentes.
    if prefixo == "auto":
        antigos = sorted(BACKUP_DIR.glob("auto_ddh_*.db"), key=lambda x: x.stat().st_mtime, reverse=True)
        for arq in antigos[30:]:
            try:
                arq.unlink()
            except OSError:
                pass
    return destino

def backup_automatico():
    """Evita criar vários backups automáticos no mesmo intervalo de 15 minutos."""
    if not Path(DB).exists():
        return None
    marcador = BACKUP_DIR / "ultimo_backup_auto.txt"
    agora = datetime.now()
    try:
        if marcador.exists():
            ultima = datetime.fromisoformat(marcador.read_text(encoding="utf-8").strip())
            if (agora - ultima).total_seconds() < 900:
                return None
    except Exception:
        pass
    destino = criar_backup("auto")
    if destino:
        marcador.write_text(agora.isoformat(), encoding="utf-8")
    return destino

def bytes_backup_atual():
    """Gera uma cópia consistente do banco para download."""
    if not Path(DB).exists():
        return b""
    with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as tmp:
        tmp_path = tmp.name
    try:
        origem = sqlite3.connect(DB, timeout=30)
        destino = sqlite3.connect(tmp_path)
        try:
            origem.backup(destino)
            destino.commit()
        finally:
            destino.close()
            origem.close()
        return Path(tmp_path).read_bytes()
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass

def validar_backup_sqlite(caminho):
    teste = sqlite3.connect(str(caminho))
    try:
        teste.execute("PRAGMA schema_version").fetchone()
        teste.execute("PRAGMA integrity_check").fetchone()
        return True
    finally:
        teste.close()

def query(sql, params=()):
    c = conn()
    try:
        return pd.read_sql_query(sql, c, params=params)
    finally:
        c.close()

def execute(sql, params=()):
    # Proteção: cria backup automático antes da primeira alteração do período.
    backup_automatico()
    c = conn()
    try:
        cur = c.cursor()
        cur.execute(sql, params)
        c.commit()
        return cur.lastrowid
    finally:
        c.close()

def delete(table, idv):
    allowed = {
        "colaboradores", "equipes", "sondas", "furos",
        "boletins", "manobras", "apontamentos", "usuarios"
    }
    if table not in allowed:
        raise ValueError("Tabela inválida.")
    execute(f"DELETE FROM {table} WHERE id=?", (int(idv),))

def abrir_boletim(bid):
    st.session_state.boletim_edit_id = int(bid)
    st.session_state.page = "📝 Novo Boletim"

def ir_para_pagina(nome):
    st.session_state.page = nome

def horas_intervalo(inicio, fim):
    a = datetime.combine(date.today(), inicio)
    b = datetime.combine(date.today(), fim)
    if b < a:
        b += timedelta(days=1)
    return round((b - a).total_seconds() / 3600, 2)

def activity_row(codigo):
    df = query("SELECT * FROM atividades WHERE codigo=?", (int(codigo),))
    return {} if df.empty else df.iloc[0].to_dict()

def entity_options(table, label_col, where=None):
    allowed = {"sondas", "equipes", "furos", "colaboradores"}
    if table not in allowed:
        raise ValueError("Tabela inválida.")
    sql = f"SELECT * FROM {table}" + (f" WHERE {where}" if where else "")
    df = query(sql)
    mapping = {} if df.empty else {
        int(r["id"]): str(r[label_col]) for _, r in df.iterrows()
    }
    return df, mapping

def safe_name(df, col, idv):
    if idv is None or pd.isna(idv) or df.empty:
        return ""
    row = df[df["id"] == int(idv)]
    return "" if row.empty else str(row.iloc[0][col])

# ============================================================
# BANCO DE DADOS
# ============================================================
def seed_activities():
    rows = [
        (1,"Segurança e Gestão","DDS / Segurança","SEGURANÇA"),
        (2,"Administrativo","Reunião / Administrativo","ADMINISTRATIVO"),
        (3,"Logística","Deslocamento / Logística","APOIO OPERACIONAL"),
        (4,"Fluidos","Preparação de fluido","APOIO OPERACIONAL"),
        (5,"Praça e Acesso","Preparação de praça / acesso","APOIO OPERACIONAL"),
        (6,"Mobilização","Mobilização / desmobilização","APOIO OPERACIONAL"),
        (7,"Manutenção Preventiva","Manutenção preventiva","MANUTENÇÃO PREVENTIVA"),
        (8,"Manutenção Corretiva","Manutenção mecânica corretiva","MECÂNICA CORRETIVA"),
        (9,"Suprimentos","Aguardar / receber suprimentos","PARADA EXTERNA"),
        (10,"Apoio Externo","Aguardar apoio externo","PARADA EXTERNA"),
        (11,"Condições Externas","Chuva / condição climática","PARADA EXTERNA"),
        (12,"Contratante","Aguardar liberação do contratante","PARADA EXTERNA"),
        (13,"Serviços Especializados","Serviço especializado","APOIO OPERACIONAL"),
        (14,"Produção","Perfuração","OPERAÇÃO DIRETA"),
        (15,"Produção","Manobra","OPERAÇÃO DIRETA"),
        (16,"Produção","Troca de haste / tubo","OPERAÇÃO DIRETA"),
        (17,"Produção","Condicionamento do furo","OPERAÇÃO DIRETA"),
        (18,"Produção","Furando e manobrando","OPERAÇÃO DIRETA"),
        (19,"Operação","Preparação operacional","APOIO OPERACIONAL"),
        (20,"Operação","Limpeza e organização","APOIO OPERACIONAL"),
        (21,"Ferramental","Troca de ferramental","APOIO OPERACIONAL"),
        (22,"Ferramental","Inspeção de ferramental","APOIO OPERACIONAL"),
        (23,"Revestimento","Instalação / retirada de revestimento","INTERVENÇÃO NO FURO"),
        (24,"Intervenção","Desvio / intervenção no furo","INTERVENÇÃO NO FURO"),
        (25,"Intervenção","Obstrução / perda no furo","INTERVENÇÃO NO FURO"),
        (26,"Segurança e Gestão","Treinamento / gestão","SEGURANÇA"),
        (27,"Administrativo","Encerramento / relatório","ADMINISTRATIVO"),
    ]
    c = conn()
    try:
        c.executemany("""
            INSERT INTO atividades(codigo,grupo,atividade,classificacao)
            VALUES(?,?,?,?)
            ON CONFLICT(codigo) DO NOTHING
        """, rows)
        c.commit()
    finally:
        c.close()

def init_db():
    c = conn()
    cur = c.cursor()

    cur.execute("""CREATE TABLE IF NOT EXISTS colaboradores(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL,
        funcao TEXT,
        matricula TEXT,
        status TEXT DEFAULT 'Ativo'
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS equipes(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo TEXT UNIQUE NOT NULL,
        nome TEXT NOT NULL,
        supervisor_id INTEGER,
        sondador_id INTEGER,
        auxiliar1_id INTEGER,
        auxiliar2_id INTEGER,
        status TEXT DEFAULT 'Ativa'
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS sondas(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo TEXT UNIQUE NOT NULL,
        modelo TEXT,
        fabricante TEXT,
        patrimonio TEXT,
        equipe_id INTEGER,
        status TEXT DEFAULT 'Operando'
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS furos(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        identificacao TEXT UNIQUE NOT NULL,
        projeto TEXT,
        cliente TEXT,
        local TEXT,
        coord_e REAL,
        coord_n REAL,
        cota REAL,
        azimute REAL,
        dip REAL,
        status TEXT DEFAULT 'Em andamento'
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS atividades(
        codigo INTEGER PRIMARY KEY,
        grupo TEXT NOT NULL,
        atividade TEXT NOT NULL,
        classificacao TEXT NOT NULL
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS boletins(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        data TEXT NOT NULL,
        turno TEXT,
        projeto TEXT,
        cliente TEXT,
        sonda_id INTEGER,
        equipe_id INTEGER,
        furo_id INTEGER,
        horimetro_inicial REAL,
        horimetro_final REAL,
        observacoes TEXT,
        criado_em TEXT
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS manobras(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        boletim_id INTEGER NOT NULL,
        numero INTEGER,
        de_m REAL,
        ate_m REAL,
        recuperado_m REAL,
        dip REAL,
        qaqc TEXT,
        perfil TEXT,
        coroa TEXT,
        revestimento TEXT,
        fluido TEXT,
        FOREIGN KEY(boletim_id) REFERENCES boletins(id) ON DELETE CASCADE
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS apontamentos(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        boletim_id INTEGER NOT NULL,
        codigo_atividade INTEGER,
        hora_inicio TEXT,
        hora_fim TEXT,
        horas REAL,
        horimetro REAL,
        observacao TEXT,
        FOREIGN KEY(boletim_id) REFERENCES boletins(id) ON DELETE CASCADE
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS usuarios(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL,
        usuario TEXT UNIQUE NOT NULL,
        senha TEXT NOT NULL,
        nivel TEXT NOT NULL,
        equipe_id INTEGER,
        status TEXT DEFAULT 'Ativo',
        criado_em TEXT
    )""")

    c.commit()
    c.close()

    seed_activities()

    # Usuário padrão, criado apenas se ainda não existir.
    if query("SELECT COUNT(*) AS total FROM usuarios WHERE usuario='admin'").iloc[0]["total"] == 0:
        execute("""
            INSERT INTO usuarios(nome,usuario,senha,nivel,status,criado_em)
            VALUES(?,?,?,?,?,?)
        """, (
            "Administrador",
            "admin",
            hash_senha("admin123"),
            "Administrador",
            "Ativo",
            datetime.now().isoformat()
        ))

init_db()

# ============================================================
# LOGO
# ============================================================
def mostrar_logo(container, largura=True):
    if LOGO_PATH.exists():
        container.image(str(LOGO_PATH), use_container_width=largura)
    else:
        container.markdown(
            "<div style='text-align:center;font-size:42px'>⛏️</div>"
            "<div style='text-align:center;font-size:28px;font-weight:800'>DDH CAMPO</div>"
            "<div style='text-align:center;letter-spacing:2px'>CONTROLE OPERACIONAL</div>",
            unsafe_allow_html=True
        )

# ============================================================
# EXCEL COM LOGO
# ============================================================
def excel_boletim(boletim_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    bd = query("SELECT * FROM boletins WHERE id=?", (int(boletim_id),))
    if bd.empty:
        return None

    b = bd.iloc[0].to_dict()
    man = query(
        "SELECT * FROM manobras WHERE boletim_id=? ORDER BY numero,id",
        (int(boletim_id),)
    )
    ap = query("""
        SELECT p.*, a.grupo, a.atividade, a.classificacao
        FROM apontamentos p
        LEFT JOIN atividades a ON a.codigo=p.codigo_atividade
        WHERE p.boletim_id=?
        ORDER BY p.id
    """, (int(boletim_id),))

    sondas = query("SELECT * FROM sondas")
    equipes = query("SELECT * FROM equipes")
    furos = query("SELECT * FROM furos")

    wb = Workbook()
    ws = wb.active
    ws.title = "BOLETIM_DDH"

    dark = PatternFill("solid", fgColor="102A43")
    gold = PatternFill("solid", fgColor="D99A00")
    light = PatternFill("solid", fgColor="EAF0F5")
    white = Font(color="FFFFFF", bold=True, size=14)
    bold = Font(bold=True)
    thin = Side(style="thin", color="A0A0A0")

    # Espaço da logo
    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 90
    if LOGO_PATH.exists():
        try:
            img = XLImage(str(LOGO_PATH))
            img.width = 145
            img.height = 145
            ws.add_image(img, "A1")
        except Exception:
            pass

    ws.merge_cells("A3:J3")
    ws["A3"] = "BOLETIM DE SONDAGEM ROTATIVA DIAMANTADA - DDH"
    ws["A3"].fill = dark
    ws["A3"].font = white
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

    sonda = safe_name(sondas, "codigo", b.get("sonda_id"))
    equipe = safe_name(equipes, "codigo", b.get("equipe_id"))
    furo = safe_name(furos, "identificacao", b.get("furo_id"))

    info = [
        ("Data", b.get("data")),
        ("Turno", b.get("turno")),
        ("Projeto", b.get("projeto")),
        ("Cliente", b.get("cliente")),
        ("Sonda", sonda),
        ("Equipe", equipe),
        ("Furo", furo),
        ("Horímetro Inicial", b.get("horimetro_inicial")),
        ("Horímetro Final", b.get("horimetro_final")),
    ]

    r = 5
    for label, value in info:
        ws.cell(r, 1, label).font = bold
        ws.cell(r, 1).fill = light
        ws.cell(r, 2, value)
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    ws.cell(r, 1, "MANOBRAS / PERFURAÇÃO").fill = dark
    ws.cell(r, 1).font = white
    ws.cell(r, 1).alignment = Alignment(horizontal="center")
    r += 1

    heads = [
        "Nº", "De (m)", "Até (m)", "Avanço (m)", "Recuperado (m)",
        "Rec. %", "DIP", "QAQC", "Perfil", "Fluido"
    ]
    for c, h in enumerate(heads, 1):
        ws.cell(r, c, h).fill = gold
        ws.cell(r, c, h).font = bold

    for _, x in man.iterrows():
        r += 1
        av = float(x["ate_m"] or 0) - float(x["de_m"] or 0)
        rec = (float(x["recuperado_m"] or 0) / av * 100) if av else 0
        vals = [
            x["numero"], x["de_m"], x["ate_m"], av, x["recuperado_m"],
            rec, x["dip"], x["qaqc"], x["perfil"], x["fluido"]
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(r, c, v)

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    ws.cell(r, 1, "ATIVIDADES / HORAS").fill = dark
    ws.cell(r, 1).font = white
    ws.cell(r, 1).alignment = Alignment(horizontal="center")
    r += 1

    heads = [
        "Código", "Grupo", "Atividade", "Classificação", "Início",
        "Fim", "Horas", "Horímetro", "Observação"
    ]
    for c, h in enumerate(heads, 1):
        ws.cell(r, c, h).fill = gold
        ws.cell(r, c, h).font = bold

    for _, x in ap.iterrows():
        r += 1
        vals = [
            x["codigo_atividade"], x["grupo"], x["atividade"],
            x["classificacao"], x["hora_inicio"], x["hora_fim"],
            x["horas"], x["horimetro"], x["observacao"]
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(r, c, v)

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    ws.cell(r, 1, f"Observações: {b.get('observacoes') or ''}")
    ws.cell(r, 1).alignment = Alignment(wrap_text=True)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = [12, 20, 28, 22, 18, 15, 12, 20, 25, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# ============================================================
# EXCEL CONSOLIDADO OPERACIONAL
# ============================================================
def excel_consolidado():
    """Gera um Excel executivo completo e recalcula horas/ROP com lógica robusta."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "DASHBOARD"

    navy = "102A43"
    blue = "1F4E78"
    gold = "D99A00"
    pale = "EAF0F5"
    green = "D9EAD3"
    red = "F4CCCC"
    dark_fill = PatternFill("solid", fgColor=navy)
    gold_fill = PatternFill("solid", fgColor=gold)
    pale_fill = PatternFill("solid", fgColor=pale)
    green_fill = PatternFill("solid", fgColor=green)
    white_bold = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=22, color=navy)
    bold = Font(bold=True)
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Produção e horas são agregadas separadamente por boletim para nunca duplicar
    # horas quando houver várias manobras e vários apontamentos no mesmo boletim.
    dados = query("""
        SELECT b.id AS boletim_id, b.data, b.turno, b.projeto, b.cliente,
               e.codigo AS equipe, e.nome AS nome_equipe,
               s.codigo AS sonda, s.modelo AS modelo_sonda,
               f.identificacao AS furo,
               COALESCE(m.metros,0) AS metros,
               COALESCE(m.recuperado,0) AS recuperado
        FROM boletins b
        LEFT JOIN equipes e ON e.id=b.equipe_id
        LEFT JOIN sondas s ON s.id=b.sonda_id
        LEFT JOIN furos f ON f.id=b.furo_id
        LEFT JOIN (
            SELECT boletim_id,
                   SUM(COALESCE(ate_m,0)-COALESCE(de_m,0)) AS metros,
                   SUM(COALESCE(recuperado_m,0)) AS recuperado
            FROM manobras GROUP BY boletim_id
        ) m ON m.boletim_id=b.id
        ORDER BY b.data,b.id
    """)

    horas = query("""
        SELECT p.boletim_id,
               COALESCE(SUM(CASE
                   WHEN (TRIM(UPPER(COALESCE(a.classificacao,'')))='OPERAÇÃO DIRETA'
                         OR p.codigo_atividade IN (14,15,16,17,18))
                   THEN COALESCE(p.horas,0) ELSE 0 END),0) AS horas_operacao,
               COALESCE(SUM(CASE
                   WHEN TRIM(UPPER(COALESCE(a.classificacao,'')))='MANUTENÇÃO PREVENTIVA'
                   THEN COALESCE(p.horas,0) ELSE 0 END),0) AS horas_manut_preventiva,
               COALESCE(SUM(CASE
                   WHEN TRIM(UPPER(COALESCE(a.classificacao,'')))='MECÂNICA CORRETIVA'
                   THEN COALESCE(p.horas,0) ELSE 0 END),0) AS horas_manut_corretiva,
               COALESCE(SUM(CASE
                   WHEN TRIM(UPPER(COALESCE(a.classificacao,'')))='PARADA EXTERNA'
                   THEN COALESCE(p.horas,0) ELSE 0 END),0) AS horas_parada_externa,
               COALESCE(SUM(COALESCE(p.horas,0)),0) AS horas_apontadas
        FROM apontamentos p
        LEFT JOIN atividades a ON a.codigo=p.codigo_atividade
        GROUP BY p.boletim_id
    """)

    base = dados.merge(horas, on="boletim_id", how="left") if not dados.empty else dados.copy()
    if base.empty:
        base = pd.DataFrame(columns=["boletim_id","data","turno","projeto","cliente","equipe","nome_equipe","sonda","modelo_sonda","furo","metros","recuperado","horas_operacao","horas_manut_preventiva","horas_manut_corretiva","horas_parada_externa","horas_apontadas"])
    numeric_cols = ["metros","recuperado","horas_operacao","horas_manut_preventiva","horas_manut_corretiva","horas_parada_externa","horas_apontadas"]
    for col in numeric_cols:
        base[col] = pd.to_numeric(base.get(col,0), errors="coerce").fillna(0.0)
    base["Recuperação %"] = (base["recuperado"] / base["metros"].replace(0,pd.NA)*100).fillna(0.0)
    base["ROP (m/h)"] = (base["metros"] / base["horas_operacao"].replace(0,pd.NA)).fillna(0.0)
    base["Horas de Parada"] = (base["horas_apontadas"] - base["horas_operacao"]).clip(lower=0)
    base["data_dt"] = pd.to_datetime(base["data"], errors="coerce")
    base["semana"] = base["data_dt"].apply(lambda x: f"{x.isocalendar().year}-S{x.isocalendar().week:02d}" if pd.notna(x) else "")
    base["mes"] = base["data_dt"].dt.strftime("%Y-%m").fillna("")

    def calc_group(df, cols):
        if df.empty:
            return pd.DataFrame(columns=cols+["Metros","Recuperado","Horas Operação","Horas Apontadas","Boletins","Recuperação %","ROP (m/h)"])
        out=df.groupby(cols,dropna=False,as_index=False).agg(
            Metros=("metros","sum"), Recuperado=("recuperado","sum"),
            **{"Horas Operação":("horas_operacao","sum"),"Horas Apontadas":("horas_apontadas","sum")},
            Boletins=("boletim_id","count"))
        out["Recuperação %"]=(out["Recuperado"]/out["Metros"].replace(0,pd.NA)*100).fillna(0)
        out["ROP (m/h)"]=(out["Metros"]/out["Horas Operação"].replace(0,pd.NA)).fillna(0)
        return out

    diaria=calc_group(base,["data"])
    semanal=calc_group(base,["semana"])
    mensal=calc_group(base,["mes"])
    equipes=calc_group(base,["equipe","nome_equipe"]).sort_values("Metros",ascending=False)
    sondas=calc_group(base,["sonda","modelo_sonda"]).sort_values("Metros",ascending=False)

    # Dashboard executivo
    ws_dash.merge_cells("A1:V1")
    ws_dash["A1"]="DDH CAMPO | DASHBOARD EXECUTIVO"
    ws_dash["A1"].font=title_font
    ws_dash["A1"].alignment=Alignment(horizontal="center",vertical="center")
    ws_dash.row_dimensions[1].height=34
    ws_dash.merge_cells("A2:V2")
    ws_dash["A2"]=f"Atualizado automaticamente em {datetime.now().strftime('%d/%m/%Y %H:%M')} | Baseado nos boletins cadastrados"
    ws_dash["A2"].alignment=Alignment(horizontal="center")

    if LOGO_PATH.exists():
        try:
            img=XLImage(str(LOGO_PATH)); img.width=78; img.height=78; ws_dash.add_image(img,"A4")
        except Exception: pass

    total_metros=float(base["metros"].sum()); total_rec=float(base["recuperado"].sum())
    total_horas=float(base["horas_operacao"].sum()); total_ap=float(base["horas_apontadas"].sum())
    total_boletins=int(len(base)); rec_pct=total_rec/total_metros*100 if total_metros else 0
    rop=total_metros/total_horas if total_horas else 0
    disponibilidade=((total_ap-base["horas_manut_corretiva"].sum())/total_ap*100) if total_ap else 0
    utilizacao=(total_horas/(total_ap-base["horas_manut_corretiva"].sum())*100) if (total_ap-base["horas_manut_corretiva"].sum())>0 else 0

    cards=[("A5","METROS TOTAIS",f"{total_metros:,.2f} m"),("D5","RECUPERAÇÃO",f"{rec_pct:.1f}%"),("G5","HORAS OPERAÇÃO",f"{total_horas:,.2f} h"),("J5","ROP MÉDIO",f"{rop:.2f} m/h"),("M5","BOLETINS",str(total_boletins)),("P5","HORAS APONTADAS",f"{total_ap:,.2f} h"),("S5","UTILIZAÇÃO",f"{utilizacao:.1f}%")]
    for cell,label,value in cards:
        c=ws_dash[cell].column; r=ws_dash[cell].row
        ws_dash.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+1)
        ws_dash.cell(r,c,label).fill=dark_fill; ws_dash.cell(r,c).font=white_bold; ws_dash.cell(r,c).alignment=Alignment(horizontal="center")
        ws_dash.merge_cells(start_row=r+1,start_column=c,end_row=r+2,end_column=c+1)
        ws_dash.cell(r+1,c,value).fill=green_fill; ws_dash.cell(r+1,c).font=Font(bold=True,size=15); ws_dash.cell(r+1,c).alignment=Alignment(horizontal="center",vertical="center")
    ws_dash["A10"]="INDICADORES DE DISPONIBILIDADE"
    ws_dash["A10"].font=bold
    ws_dash["D10"]="Disponibilidade Física"; ws_dash["E10"]=f"{disponibilidade:.1f}%"
    ws_dash["G10"]="Horas de Parada"; ws_dash["H10"]=f"{max(total_ap-total_horas,0):.2f} h"

    def write_df(name,df,title):
        ws=wb.create_sheet(name)
        n=max(1,len(df.columns))
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=n)
        ws["A1"]=title; ws["A1"].fill=dark_fill; ws["A1"].font=white_bold; ws["A1"].alignment=Alignment(horizontal="center")
        headers=list(df.columns)
        for i,h in enumerate(headers,1):
            cell=ws.cell(3,i,h); cell.fill=gold_fill; cell.font=bold; cell.alignment=Alignment(horizontal="center")
        for ri,row in enumerate(df.itertuples(index=False,name=None),4):
            for ci,val in enumerate(row,1): ws.cell(ri,ci,None if pd.isna(val) else val)
        for row in ws.iter_rows(min_row=3,max_row=max(3,len(df)+3),min_col=1,max_col=n):
            for cell in row: cell.border=border; cell.alignment=Alignment(vertical="center")
        for i,h in enumerate(headers,1): ws.column_dimensions[get_column_letter(i)].width=max(13,min(28,len(str(h))+5))
        ws.freeze_panes="A4"; ws.auto_filter.ref=f"A3:{get_column_letter(n)}{max(3,len(df)+3)}"; ws.sheet_view.showGridLines=False
        if len(df)>0:
            tab=Table(displayName="Tbl"+re.sub(r"[^A-Za-z0-9]","",name.title()),ref=f"A3:{get_column_letter(n)}{len(df)+3}")
            tab.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True,showColumnStripes=False)
            ws.add_table(tab)
        # Destaque de indicadores
        for idx,h in enumerate(headers,1):
            if h in ("Metros","metros"):
                ws.conditional_formatting.add(f"{get_column_letter(idx)}4:{get_column_letter(idx)}{max(4,len(df)+3)}",DataBarRule(start_type="min", end_type="max", color="4472C4", showValue=True))
            if h in ("ROP (m/h)","Recuperação %"):
                ws.conditional_formatting.add(f"{get_column_letter(idx)}4:{get_column_letter(idx)}{max(4,len(df)+3)}",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="percentile",mid_value=50,mid_color="FFEB84",end_type="max",end_color="63BE7B"))
        return ws

    base_export=base.drop(columns=["data_dt"],errors="ignore")
    ws_base=write_df("DADOS_BOLETINS",base_export,"DADOS CONSOLIDADOS DOS BOLETINS")
    ws_diaria=write_df("PRODUCAO_DIARIA",diaria,"PRODUÇÃO DIÁRIA")
    ws_sem=write_df("PRODUCAO_SEMANAL",semanal,"PRODUÇÃO SEMANAL")
    ws_mes=write_df("PRODUCAO_MENSAL",mensal,"PRODUÇÃO MENSAL")
    ws_eq=write_df("EQUIPES",equipes,"RANKING E PRODUÇÃO POR EQUIPE")
    ws_s=write_df("SONDAS",sondas,"RANKING E PRODUÇÃO POR SONDA")

    # Evolução acumulada
    evolucao=diaria.copy()
    if not evolucao.empty:
        evolucao["Metros Acumulados"]=evolucao["Metros"].cumsum()
    ws_ev=write_df("EVOLUCAO_ACUMULADA",evolucao,"EVOLUÇÃO DA PRODUÇÃO ACUMULADA")

    # Resumo das classificações de tempo
    tempo=query("""
        SELECT COALESCE(NULLIF(TRIM(a.classificacao),''),'SEM CLASSIFICAÇÃO') AS Classificação,
               SUM(COALESCE(p.horas,0)) AS Horas
        FROM apontamentos p LEFT JOIN atividades a ON a.codigo=p.codigo_atividade
        GROUP BY COALESCE(NULLIF(TRIM(a.classificacao),''),'SEM CLASSIFICAÇÃO')
        ORDER BY Horas DESC
    """)
    ws_t=write_df("HORAS_CLASSIFICACAO",tempo,"HORAS POR CLASSIFICAÇÃO OPERACIONAL")

    # Gráficos executivos
    if len(equipes)>0:
        ch=BarChart(); ch.title="Produção por Equipe"; ch.y_axis.title="Metros"; ch.x_axis.title="Equipe"; ch.height=8; ch.width=14
        ch.add_data(Reference(ws_eq,min_col=3,min_row=3,max_row=len(equipes)+3),titles_from_data=True)
        ch.set_categories(Reference(ws_eq,min_col=1,min_row=4,max_row=len(equipes)+3)); ws_dash.add_chart(ch,"A13")
    if len(evolucao)>0:
        ch=LineChart(); ch.title="Produção Acumulada"; ch.y_axis.title="Metros"; ch.height=8; ch.width=14
        col=evolucao.columns.get_loc("Metros Acumulados")+1
        ch.add_data(Reference(ws_ev,min_col=col,min_row=3,max_row=len(evolucao)+3),titles_from_data=True)
        ch.set_categories(Reference(ws_ev,min_col=1,min_row=4,max_row=len(evolucao)+3)); ws_dash.add_chart(ch,"P13")
    if len(tempo)>0:
        ch=PieChart(); ch.title="Distribuição das Horas"; ch.height=8; ch.width=12
        ch.add_data(Reference(ws_t,min_col=2,min_row=3,max_row=len(tempo)+3),titles_from_data=True)
        ch.set_categories(Reference(ws_t,min_col=1,min_row=4,max_row=len(tempo)+3)); ws_dash.add_chart(ch,"A30")
    if len(sondas)>0:
        ch=BarChart(); ch.title="ROP por Sonda"; ch.y_axis.title="m/h"; ch.height=8; ch.width=14
        col=sondas.columns.get_loc("ROP (m/h)")+1
        ch.add_data(Reference(ws_s,min_col=col,min_row=3,max_row=len(sondas)+3),titles_from_data=True)
        ch.set_categories(Reference(ws_s,min_col=1,min_row=4,max_row=len(sondas)+3)); ws_dash.add_chart(ch,"P30")

    for col in range(1,23): ws_dash.column_dimensions[get_column_letter(col)].width=14
    ws_dash.sheet_view.showGridLines=False
    out=BytesIO(); wb.save(out); return out.getvalue()

# ============================================================
# LOGIN
# ============================================================
def tela_login():
    left, center, right = st.columns([1, 1.2, 1])

    with center:
        mostrar_logo(st)
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("### 🔐 Acesso ao Sistema")
        st.caption("Entre com seu usuário e senha para acessar o DDH CAMPO.")

        with st.form("login_form"):
            usuario = st.text_input("Usuário")
            senha = st.text_input("Senha", type="password")
            entrar = st.form_submit_button(
                "🔐 ENTRAR",
                type="primary",
                use_container_width=True
            )

        if entrar:
            df = query("""
                SELECT *
                FROM usuarios
                WHERE usuario=? AND status='Ativo'
            """, (usuario.strip(),))

            if df.empty:
                st.error("❌ Usuário não encontrado ou inativo.")
            else:
                u = df.iloc[0]
                if verificar_senha(senha, u["senha"]):
                    st.session_state.logado = True
                    st.session_state.usuario_id = int(u["id"])
                    st.session_state.usuario_nome = str(u["nome"])
                    st.session_state.usuario_login = str(u["usuario"])
                    st.session_state.usuario_nivel = str(u["nivel"])

                    if u["nivel"] == "Campo":
                        st.session_state.page = "📝 Novo Boletim"
                    else:
                        st.session_state.page = "🏠 Painel DDH"

                    st.rerun()
                else:
                    st.error("❌ Senha incorreta.")

        st.info("Usuário inicial: **admin** | Senha inicial: **admin123**")

if not st.session_state.logado:
    tela_login()
    st.stop()

# ============================================================
# MENU
# ============================================================
nivel = st.session_state.usuario_nivel

if nivel == "Administrador":
    PAGES = PAGES_ADMIN
elif nivel == "Supervisor":
    PAGES = PAGES_SUPERVISOR
else:
    PAGES = PAGES_CAMPO

if st.session_state.page not in PAGES:
    st.session_state.page = PAGES[0]

with st.sidebar:
    mostrar_logo(st)
    st.divider()
    st.caption(f"👤 {st.session_state.usuario_nome}")
    st.caption(f"🔐 Perfil: {nivel}")

    idx_pagina = PAGES.index(st.session_state.page)

    def ao_mudar_menu():
        st.session_state.page = st.session_state.menu_opcao

    st.radio(
        "Menu",
        PAGES,
        index=idx_pagina,
        key="menu_opcao",
        on_change=ao_mudar_menu
    )

    st.divider()

    if st.button("🚪 Sair do Sistema", use_container_width=True):
        for k in ["logado", "usuario_id", "usuario_nome", "usuario_login", "usuario_nivel"]:
            st.session_state[k] = DEFAULT_SESSION[k]
        st.session_state.page = "🏠 Painel DDH"
        st.session_state.boletim_edit_id = None
        st.rerun()

page = st.session_state.page

# ============================================================
# PAINEL
# ============================================================
if page == "🏠 Painel DDH":
    st.title("🏠 PAINEL DDH")

    resumo = query("""
        SELECT
            COALESCE(SUM(COALESCE(ate_m,0)-COALESCE(de_m,0)),0) metros,
            COALESCE(SUM(COALESCE(recuperado_m,0)),0) recuperado
        FROM manobras
    """)

    horas = query("""
        SELECT COALESCE(SUM(
            CASE WHEN (TRIM(UPPER(COALESCE(a.classificacao,'')))='OPERAÇÃO DIRETA' OR p.codigo_atividade IN (14,15,16,17,18))
            THEN COALESCE(p.horas,0) ELSE 0 END
        ),0) horas_operacao
        FROM apontamentos p
        LEFT JOIN atividades a ON a.codigo=p.codigo_atividade
    """)

    metros = float(resumo.iloc[0]["metros"] or 0)
    recuperado = float(resumo.iloc[0]["recuperado"] or 0)
    horas_total = float(horas.iloc[0]["horas_operacao"] or 0)
    rec_total = recuperado / metros * 100 if metros else 0
    rop = metros / horas_total if horas_total else 0
    nboletins = int(query("SELECT COUNT(*) total FROM boletins").iloc[0]["total"])

    cols = st.columns(5)
    cols[0].metric("⛏️ PRODUÇÃO TOTAL", f"{metros:.2f} m")
    cols[1].metric("🧪 RECUPERAÇÃO", f"{rec_total:.1f}%")
    cols[2].metric("⏱️ HORAS OPERAÇÃO", f"{horas_total:.2f} h")
    cols[3].metric("⚡ ROP MÉDIO", f"{rop:.2f} m/h")
    cols[4].metric("📋 BOLETINS", nboletins)

    st.divider()
    st.subheader("👥 Produção por Equipe")

    equipes_df = query("""
        SELECT e.id equipe_id, e.codigo equipe, e.nome nome_equipe,
               COALESCE(m.metros,0) metros,
               COALESCE(m.recuperado,0) recuperado,
               COALESCE(h.horas_operacao,0) horas_operacao
        FROM equipes e
        LEFT JOIN (
            SELECT b.equipe_id,
                   SUM(COALESCE(m.ate_m,0)-COALESCE(m.de_m,0)) metros,
                   SUM(COALESCE(m.recuperado_m,0)) recuperado
            FROM boletins b
            LEFT JOIN manobras m ON m.boletim_id=b.id
            GROUP BY b.equipe_id
        ) m ON m.equipe_id=e.id
        LEFT JOIN (
            SELECT b.equipe_id,
                   SUM(CASE WHEN (TRIM(UPPER(COALESCE(a.classificacao,'')))='OPERAÇÃO DIRETA' OR p.codigo_atividade IN (14,15,16,17,18))
                       THEN COALESCE(p.horas,0) ELSE 0 END) horas_operacao
            FROM boletins b
            LEFT JOIN apontamentos p ON p.boletim_id=b.id
            LEFT JOIN atividades a ON a.codigo=p.codigo_atividade
            GROUP BY b.equipe_id
        ) h ON h.equipe_id=e.id
        WHERE e.status!='Inativa'
        ORDER BY metros DESC
    """)

    if not equipes_df.empty:
        equipes_df["Recuperação %"] = (
            equipes_df["recuperado"] /
            equipes_df["metros"].replace(0, pd.NA) * 100
        ).fillna(0)
        equipes_df["ROP (m/h)"] = (
            equipes_df["metros"] /
            equipes_df["horas_operacao"].replace(0, pd.NA)
        ).fillna(0)

        show = equipes_df[
            ["equipe","nome_equipe","metros","recuperado",
             "Recuperação %","horas_operacao","ROP (m/h)"]
        ].copy()
        show.columns = [
            "Equipe","Nome","Metros","Recuperado",
            "Recuperação %","Horas Operação","ROP (m/h)"
        ]
        st.dataframe(show.round(2), use_container_width=True, hide_index=True)

    st.divider()
    st.subheader("🔩 Produção por Sonda")

    sondas_df = query("""
        SELECT s.id sonda_id, s.codigo sonda, s.modelo, s.status,
               e.codigo equipe, e.nome nome_equipe,
               COALESCE(m.metros,0) metros,
               COALESCE(m.recuperado,0) recuperado,
               COALESCE(h.horas_operacao,0) horas_operacao
        FROM sondas s
        LEFT JOIN equipes e ON e.id=s.equipe_id
        LEFT JOIN (
            SELECT b.sonda_id,
                   SUM(COALESCE(m.ate_m,0)-COALESCE(m.de_m,0)) metros,
                   SUM(COALESCE(m.recuperado_m,0)) recuperado
            FROM boletins b
            LEFT JOIN manobras m ON m.boletim_id=b.id
            GROUP BY b.sonda_id
        ) m ON m.sonda_id=s.id
        LEFT JOIN (
            SELECT b.sonda_id,
                   SUM(CASE WHEN (TRIM(UPPER(COALESCE(a.classificacao,'')))='OPERAÇÃO DIRETA' OR p.codigo_atividade IN (14,15,16,17,18))
                       THEN COALESCE(p.horas,0) ELSE 0 END) horas_operacao
            FROM boletins b
            LEFT JOIN apontamentos p ON p.boletim_id=b.id
            LEFT JOIN atividades a ON a.codigo=p.codigo_atividade
            GROUP BY b.sonda_id
        ) h ON h.sonda_id=s.id
        WHERE s.status!='Inativa'
        ORDER BY metros DESC
    """)

    if not sondas_df.empty:
        sondas_df["Recuperação %"] = (
            sondas_df["recuperado"] /
            sondas_df["metros"].replace(0, pd.NA) * 100
        ).fillna(0)
        sondas_df["ROP (m/h)"] = (
            sondas_df["metros"] /
            sondas_df["horas_operacao"].replace(0, pd.NA)
        ).fillna(0)

        final = sondas_df[
            ["sonda","modelo","equipe","nome_equipe","metros",
             "recuperado","Recuperação %","horas_operacao","ROP (m/h)"]
        ].copy()
        final.columns = [
            "Sonda","Modelo","Equipe","Nome da Equipe","Metros",
            "Recuperado","Recuperação %","Horas Operação","ROP (m/h)"
        ]

        st.dataframe(final.round(2), use_container_width=True, hide_index=True)

        if not final.empty:
            st.subheader("📊 Produção das Sondas")
            st.bar_chart(final[["Sonda","Metros"]].set_index("Sonda"))
            st.subheader("⚡ ROP por Sonda")
            st.bar_chart(final[["Sonda","ROP (m/h)"]].set_index("Sonda"))

# ============================================================
# NOVO / EDITAR BOLETIM
# ============================================================
elif page == "📝 Novo Boletim":
    st.title("📝 NOVO RDO / BOLETIM DDH")

    df_s, map_s = entity_options("sondas", "codigo", "status != 'Inativa'")
    df_e, map_e = entity_options("equipes", "codigo", "status != 'Inativa'")
    df_f, map_f = entity_options("furos", "identificacao")

    if df_s.empty or df_e.empty or df_f.empty:
        st.warning("Cadastre pelo menos uma Sonda, uma Equipe e um Furo.")
        st.stop()

    if st.session_state.boletim_edit_id is None:
        with st.form("cabecalho"):
            a, b, c, d = st.columns(4)
            data_b = a.date_input("Data", value=date.today())
            turno = b.selectbox("Turno", ["Diurno","Noturno"])
            sonda_id = c.selectbox(
                "Sonda", list(map_s), format_func=lambda x: map_s[x]
            )

            row_s = df_s[df_s["id"] == sonda_id]
            equipe_padrao = (
                int(row_s.iloc[0]["equipe_id"])
                if not row_s.empty and pd.notna(row_s.iloc[0]["equipe_id"])
                else None
            )

            ids_e = list(map_e)
            idx = ids_e.index(equipe_padrao) if equipe_padrao in ids_e else 0
            equipe_id = d.selectbox(
                "Equipe", ids_e, index=idx,
                format_func=lambda x: map_e[x]
            )

            a, b, c, d = st.columns(4)
            furo_id = a.selectbox(
                "Furo", list(map_f), format_func=lambda x: map_f[x]
            )
            projeto = b.text_input("Projeto")
            cliente = c.text_input("Cliente")
            h_ini = d.number_input("Horímetro Inicial", min_value=0.0, step=0.1)
            h_fim = st.number_input("Horímetro Final", min_value=0.0, step=0.1)
            observacoes = st.text_area("Observações gerais")

            salvar = st.form_submit_button("💾 Criar Boletim", type="primary")

        if salvar:
            if h_fim < h_ini:
                st.error("Horímetro final não pode ser menor que o inicial.")
            else:
                bid = execute("""
                    INSERT INTO boletins(
                        data,turno,projeto,cliente,sonda_id,equipe_id,furo_id,
                        horimetro_inicial,horimetro_final,observacoes,criado_em
                    )
                    VALUES(?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    str(data_b), turno, projeto, cliente,
                    int(sonda_id), int(equipe_id), int(furo_id),
                    h_ini, h_fim, observacoes, datetime.now().isoformat()
                ))
                abrir_boletim(bid)
                st.success("Boletim criado com sucesso.")
                st.rerun()

    bid = st.session_state.boletim_edit_id

    if bid:
        bd = query("""
            SELECT b.*, s.codigo sonda, e.codigo equipe, f.identificacao furo
            FROM boletins b
            LEFT JOIN sondas s ON s.id=b.sonda_id
            LEFT JOIN equipes e ON e.id=b.equipe_id
            LEFT JOIN furos f ON f.id=b.furo_id
            WHERE b.id=?
        """, (int(bid),))

        if bd.empty:
            st.session_state.boletim_edit_id = None
            st.warning("Boletim não encontrado.")
            st.rerun()

        st.divider()
        st.subheader(f"📋 Lançamentos do Boletim #{bid}")

        with st.expander("ℹ️ Dados do Boletim", expanded=True):
            st.dataframe(bd, use_container_width=True, hide_index=True)

        st.subheader("⛏️ Manobras e Perfuração")
        dfm = query(
            "SELECT * FROM manobras WHERE boletim_id=? ORDER BY numero,id",
            (bid,)
        )
        proximo_numero = 1 if dfm.empty else int(dfm["numero"].max()) + 1

        with st.form("nova_manobra", clear_on_submit=True):
            c1, c2, c3, c4 = st.columns(4)
            numero = c1.number_input(
                "Nº", min_value=1, value=proximo_numero, step=1
            )
            de_m = c2.number_input("De (m)", min_value=0.0, step=0.1)
            ate_m = c3.number_input("Até (m)", min_value=0.0, step=0.1)
            recuperado = c4.number_input(
                "Recuperado (m)", min_value=0.0, step=0.01
            )

            av = max(0.0, ate_m - de_m)
            rec = recuperado / av * 100 if av else 0
            st.info(f"📏 Avanço: {av:.2f} m | 🧪 Recuperação: {rec:.1f}%")

            c1, c2, c3, c4, c5 = st.columns(5)
            dip = c1.number_input("DIP", step=0.1)
            qaqc = c2.text_input("QAQC")
            perfil = c3.text_input("Perfil / Diâmetro")
            coroa = c4.text_input("Coroa / Série")
            revestimento = c5.text_input("Revestimento")
            fluido = st.text_input("Tipo de Fluido")

            add = st.form_submit_button("➕ Adicionar Manobra")

        if add:
            if ate_m <= de_m:
                st.error("O valor 'Até' deve ser maior que 'De'.")
            elif recuperado > av:
                st.error("A recuperação não pode ser maior que o avanço.")
            else:
                execute("""
                    INSERT INTO manobras(
                        boletim_id,numero,de_m,ate_m,recuperado_m,dip,
                        qaqc,perfil,coroa,revestimento,fluido
                    )
                    VALUES(?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    bid, int(numero), de_m, ate_m, recuperado, dip,
                    qaqc, perfil, coroa, revestimento, fluido
                ))
                st.success("Manobra adicionada.")
                st.rerun()

        if not dfm.empty:
            view = dfm.copy()
            view["Avanço"] = view["ate_m"] - view["de_m"]
            view["Recuperação %"] = (
                view["recuperado_m"] /
                view["Avanço"].replace(0, pd.NA) * 100
            ).fillna(0)
            st.dataframe(view.round(2), use_container_width=True, hide_index=True)

            excluir_manobra = st.selectbox(
                "Excluir manobra",
                view["id"].tolist(),
                format_func=lambda x: f"Manobra #{int(view[view.id==x].iloc[0]['numero'])}"
            )
            if st.button("🗑️ Excluir Manobra"):
                delete("manobras", excluir_manobra)
                st.rerun()

        st.divider()
        st.subheader("⏱️ Atividades e Horários")

        acts = query("SELECT * FROM atividades ORDER BY codigo")
        codes = acts["codigo"].tolist()

        with st.form("nova_atividade", clear_on_submit=True):
            codigo = st.selectbox("Código da atividade", codes, key="nova_atividade_codigo")
            ar = activity_row(codigo)

            c1, c2, c3 = st.columns(3)
            c1.text_input("Grupo", value=ar.get("grupo",""), disabled=True)
            c2.text_input("Atividade", value=ar.get("atividade",""), disabled=True)
            c3.text_input(
                "Classificação",
                value=ar.get("classificacao",""),
                disabled=True
            )

            c1, c2, c3, c4 = st.columns(4)
            inicio = c1.time_input("Hora inicial", value=time(7,0))
            fim = c2.time_input("Hora final", value=time(8,0))
            horas = horas_intervalo(inicio, fim)
            c3.metric("Tempo automático", f"{horas:.2f} h")
            horimetro = c4.number_input("Horímetro", min_value=0.0, step=0.1)

            obs = st.text_input("Observação")
            add_a = st.form_submit_button("➕ Adicionar Atividade")

        if add_a:
            execute("""
                INSERT INTO apontamentos(
                    boletim_id,codigo_atividade,hora_inicio,hora_fim,
                    horas,horimetro,observacao
                )
                VALUES(?,?,?,?,?,?,?)
            """, (
                bid, int(codigo),
                inicio.strftime("%H:%M"),
                fim.strftime("%H:%M"),
                horas, horimetro, obs
            ))
            st.success("Atividade adicionada.")
            st.rerun()

        dfa = query("""
            SELECT p.id,p.codigo_atividade,a.grupo,a.atividade,
                   a.classificacao,p.hora_inicio,p.hora_fim,p.horas,
                   p.horimetro,p.observacao
            FROM apontamentos p
            LEFT JOIN atividades a ON a.codigo=p.codigo_atividade
            WHERE p.boletim_id=?
            ORDER BY p.id
        """, (bid,))

        if not dfa.empty:
            st.dataframe(dfa, use_container_width=True, hide_index=True)
            excluir_ap = st.selectbox(
                "Excluir atividade",
                dfa["id"].tolist(),
                format_func=lambda x: (
                    f"{dfa[dfa.id==x].iloc[0]['codigo_atividade']} - "
                    f"{dfa[dfa.id==x].iloc[0]['atividade']}"
                )
            )
            if st.button("🗑️ Excluir Atividade"):
                delete("apontamentos", excluir_ap)
                st.rerun()

        st.divider()
        c1, c2 = st.columns(2)

        c1.download_button(
            "📥 Baixar Excel deste Boletim",
            data=excel_boletim(bid),
            file_name=f"BOLETIM_DDH_{bid}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )

        if c2.button("➕ Novo Boletim", use_container_width=True):
            st.session_state.boletim_edit_id = None
            st.rerun()

# ============================================================
# PRODUÇÃO DIÁRIA
# ============================================================
elif page == "📅 Produção Diária":
    st.title("📅 PRODUÇÃO DIÁRIA")
    st.caption("A produção é atualizada automaticamente conforme as manobras são lançadas nos boletins.")

    datas = query("SELECT DISTINCT data FROM boletins ORDER BY data DESC")
    if datas.empty:
        st.info("Nenhuma produção registrada ainda.")
    else:
        opcoes_datas = pd.to_datetime(datas["data"], errors="coerce").dt.date.dropna().tolist()
        data_padrao = date.today() if date.today() in opcoes_datas else opcoes_datas[0]
        data_filtro = st.date_input("📅 Selecione a data", value=data_padrao, key="producao_diaria_data")

        diaria = query("""
            SELECT
                b.data,
                e.codigo AS equipe,
                e.nome AS nome_equipe,
                s.codigo AS sonda,
                f.identificacao AS furo,
                COALESCE(SUM(COALESCE(m.ate_m,0) - COALESCE(m.de_m,0)),0) AS metros,
                COALESCE(SUM(COALESCE(m.recuperado_m,0)),0) AS recuperado
            FROM boletins b
            LEFT JOIN equipes e ON e.id=b.equipe_id
            LEFT JOIN sondas s ON s.id=b.sonda_id
            LEFT JOIN furos f ON f.id=b.furo_id
            LEFT JOIN manobras m ON m.boletim_id=b.id
            WHERE b.data=?
            GROUP BY b.id, b.data, e.codigo, e.nome, s.codigo, f.identificacao
            ORDER BY metros DESC, e.codigo
        """, (str(data_filtro),))

        horas = query("""
            SELECT b.id AS boletim_id,
                   COALESCE(SUM(CASE WHEN (TRIM(UPPER(COALESCE(a.classificacao,'')))='OPERAÇÃO DIRETA' OR p.codigo_atividade IN (14,15,16,17,18))
                       THEN COALESCE(p.horas,0) ELSE 0 END),0) AS horas_operacao
            FROM boletins b
            LEFT JOIN apontamentos p ON p.boletim_id=b.id
            LEFT JOIN atividades a ON a.codigo=p.codigo_atividade
            WHERE b.data=?
            GROUP BY b.id
        """, (str(data_filtro),))

        boletim_ids = query("SELECT id FROM boletins WHERE data=?", (str(data_filtro),))
        if diaria.empty:
            st.warning("Nenhuma manobra registrada para esta data.")
        else:
            # Horas agregadas por equipe/sonda/furo não são necessárias para a produção em metros.
            diaria["Recuperação %"] = (diaria["recuperado"] / diaria["metros"].replace(0, pd.NA) * 100).fillna(0)

            total_metros = float(diaria["metros"].sum())
            total_recuperado = float(diaria["recuperado"].sum())
            total_equipes = int(diaria["equipe"].nunique())
            total_sondas = int(diaria["sonda"].nunique())
            rec_total = total_recuperado / total_metros * 100 if total_metros else 0

            c1,c2,c3,c4 = st.columns(4)
            c1.metric("⛏️ METROS DO DIA", f"{total_metros:.2f} m")
            c2.metric("👥 EQUIPES PRODUZINDO", total_equipes)
            c3.metric("🔩 SONDAS ATIVAS", total_sondas)
            c4.metric("🧪 RECUPERAÇÃO MÉDIA", f"{rec_total:.1f}%")

            st.divider()
            st.subheader("👥 Produção por Equipe")
            por_equipe = diaria.groupby(["equipe","nome_equipe"], dropna=False, as_index=False).agg(
                Metros=("metros","sum"), Recuperado=("recuperado","sum")
            )
            por_equipe["Recuperação %"] = (por_equipe["Recuperado"] / por_equipe["Metros"].replace(0,pd.NA)*100).fillna(0)
            por_equipe = por_equipe.sort_values("Metros", ascending=False)
            st.dataframe(por_equipe.round(2), use_container_width=True, hide_index=True)
            st.bar_chart(por_equipe.set_index("equipe")[["Metros"]])

            st.divider()
            st.subheader("📋 Detalhamento dos Boletins do Dia")
            detalhe = diaria[["data","equipe","nome_equipe","sonda","furo","metros","recuperado","Recuperação %"]].copy()
            detalhe.columns = ["Data","Equipe","Nome da Equipe","Sonda","Furo","Metros","Recuperado","Recuperação %"]
            st.dataframe(detalhe.round(2), use_container_width=True, hide_index=True)

            st.divider()
            st.subheader("📈 Produção diária recente")
            historico = query("""
                SELECT b.data, e.codigo AS equipe,
                       COALESCE(SUM(COALESCE(m.ate_m,0)-COALESCE(m.de_m,0)),0) AS metros
                FROM boletins b
                LEFT JOIN equipes e ON e.id=b.equipe_id
                LEFT JOIN manobras m ON m.boletim_id=b.id
                GROUP BY b.data, e.codigo
                ORDER BY b.data DESC
                LIMIT 500
            """)
            if not historico.empty:
                historico["data"] = pd.to_datetime(historico["data"], errors="coerce")
                grafico = historico.pivot_table(index="data", columns="equipe", values="metros", aggfunc="sum").fillna(0).sort_index()
                st.line_chart(grafico)

# ============================================================
# RELATÓRIOS E EXCEL CONSOLIDADO
# ============================================================
elif page == "📊 Relatórios e Excel":
    st.title("📊 RELATÓRIOS E EXCEL CONSOLIDADO")
    st.caption("Acompanhe os dados e gere uma planilha completa para apresentações diárias, semanais e mensais.")

    base_rel = query("""
        SELECT b.data,e.codigo AS equipe,s.codigo AS sonda,
               COALESCE(SUM(COALESCE(m.ate_m,0)-COALESCE(m.de_m,0)),0) AS metros,
               COALESCE(SUM(COALESCE(m.recuperado_m,0)),0) AS recuperado
        FROM boletins b
        LEFT JOIN equipes e ON e.id=b.equipe_id
        LEFT JOIN sondas s ON s.id=b.sonda_id
        LEFT JOIN manobras m ON m.boletim_id=b.id
        GROUP BY b.id,b.data,e.codigo,s.codigo
        ORDER BY b.data DESC
    """)

    if base_rel.empty:
        st.info("Ainda não existem boletins com produção para gerar o consolidado.")
    else:
        base_rel["Recuperação %"]=(base_rel["recuperado"]/base_rel["metros"].replace(0,pd.NA)*100).fillna(0)
        base_rel["data_dt"]=pd.to_datetime(base_rel["data"],errors="coerce")
        base_rel["semana"]=base_rel["data_dt"].apply(lambda x: f"{x.isocalendar().year}-S{x.isocalendar().week:02d}" if pd.notna(x) else "")
        base_rel["mes"]=base_rel["data_dt"].dt.strftime("%Y-%m").fillna("")

        total=float(base_rel["metros"].sum())
        rec=float(base_rel["recuperado"].sum())
        c1,c2,c3=st.columns(3)
        c1.metric("⛏️ METROS CONSOLIDADOS",f"{total:.2f} m")
        c2.metric("🧪 RECUPERAÇÃO MÉDIA",f"{(rec/total*100 if total else 0):.1f}%")
        c3.metric("📋 LANÇAMENTOS",len(base_rel))

        st.subheader("📅 Produção por Semana")
        semanal_view=base_rel.groupby("semana",as_index=False)["metros"].sum().sort_values("semana")
        st.dataframe(semanal_view,use_container_width=True,hide_index=True)
        st.bar_chart(semanal_view.set_index("semana"))

        st.subheader("📆 Produção por Mês")
        mensal_view=base_rel.groupby("mes",as_index=False)["metros"].sum().sort_values("mes")
        st.dataframe(mensal_view,use_container_width=True,hide_index=True)
        st.bar_chart(mensal_view.set_index("mes"))

        st.divider()
        st.success("O Excel consolidado reúne Dashboard, dados dos boletins, produção diária, semanal, mensal, equipes e sondas.")
        st.download_button(
            "📥 BAIXAR EXCEL CONSOLIDADO COMPLETO",
            data=excel_consolidado(),
            file_name=f"DDH_CAMPO_CONSOLIDADO_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
            key="baixar_excel_consolidado"
        )

# ============================================================
# BOLETINS SALVOS
# ============================================================
elif page == "📋 Boletins Salvos":
    st.title("📋 BOLETINS SALVOS")

    df = query("""
        SELECT b.*, s.codigo sonda, e.codigo equipe, f.identificacao furo
        FROM boletins b
        LEFT JOIN sondas s ON s.id=b.sonda_id
        LEFT JOIN equipes e ON e.id=b.equipe_id
        LEFT JOIN furos f ON f.id=b.furo_id
        ORDER BY b.data DESC,b.id DESC
    """)

    if df.empty:
        st.info("Nenhum boletim salvo.")
    else:
        st.dataframe(df, use_container_width=True, hide_index=True)

        bid = st.selectbox(
            "Selecione um boletim",
            df["id"].tolist(),
            format_func=lambda x: (
                f"Boletim #{x} | "
                f"{df[df.id==x].iloc[0]['data']} | "
                f"{df[df.id==x].iloc[0]['furo']}"
            )
        )

        c1, c2, c3 = st.columns(3)

        c1.button(
            "✏️ Abrir para lançamento",
            on_click=abrir_boletim,
            args=(int(bid),),
            use_container_width=True
        )

        c2.download_button(
            "📥 Baixar Excel",
            data=excel_boletim(int(bid)),
            file_name=f"BOLETIM_DDH_{bid}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

        if c3.button("🗑️ Excluir boletim", use_container_width=True):
            execute("DELETE FROM manobras WHERE boletim_id=?", (int(bid),))
            execute("DELETE FROM apontamentos WHERE boletim_id=?", (int(bid),))
            delete("boletins", bid)
            if st.session_state.boletim_edit_id == int(bid):
                st.session_state.boletim_edit_id = None
            st.success("Boletim excluído.")
            st.rerun()

# ============================================================
# COLABORADORES
# ============================================================
elif page == "👷 Colaboradores":
    st.title("👷 COLABORADORES")
    t1, t2 = st.tabs(["Lista", "Novo cadastro"])

    with t1:
        df = query("SELECT * FROM colaboradores ORDER BY nome")
        st.dataframe(df, use_container_width=True, hide_index=True)

        if not df.empty:
            idx = st.selectbox(
                "Excluir colaborador",
                df["id"].tolist(),
                format_func=lambda x: (
                    f"{x} - {df[df.id==x].iloc[0]['nome']}"
                )
            )

            if st.button("🗑️ Excluir colaborador"):
                refs = query("""
                    SELECT
                    (SELECT COUNT(*) FROM equipes WHERE supervisor_id=? OR sondador_id=? OR auxiliar1_id=? OR auxiliar2_id=?) total
                """, (idx, idx, idx, idx))
                if int(refs.iloc[0]["total"]) > 0:
                    st.error("Não é possível excluir: colaborador vinculado a uma equipe.")
                else:
                    delete("colaboradores", idx)
                    st.rerun()

    with t2:
        with st.form("form_colab", clear_on_submit=True):
            nome = st.text_input("Nome")
            funcao = st.selectbox(
                "Função",
                [
                    "Supervisor","Sondador","Auxiliar de Sondador",
                    "Geólogo","Mecânico","Técnico de Segurança","Outro"
                ]
            )
            matricula = st.text_input("Matrícula")
            status = st.selectbox("Status", ["Ativo","Inativo"])
            salvar = st.form_submit_button("Cadastrar", type="primary")

        if salvar:
            if nome.strip():
                execute("""
                    INSERT INTO colaboradores(nome,funcao,matricula,status)
                    VALUES(?,?,?,?)
                """, (nome.strip(), funcao, matricula, status))
                st.success("Colaborador cadastrado.")
                st.rerun()
            else:
                st.error("Informe o nome.")

# ============================================================
# EQUIPES
# ============================================================
elif page == "👥 Equipes":
    st.title("👥 EQUIPES")
    dfc = query("SELECT * FROM colaboradores WHERE status='Ativo' ORDER BY nome")

    if dfc.empty:
        st.warning("Cadastre colaboradores primeiro.")
    else:
        opts = dfc["id"].tolist()

        def fmt_colaborador(x):
            row = dfc[dfc.id == x]
            return str(x) if row.empty else str(row.iloc[0]["nome"])

        with st.form("form_equipe", clear_on_submit=True):
            c1, c2 = st.columns(2)
            codigo = c1.text_input("Código da equipe")
            nome = c2.text_input("Nome da equipe")
            supervisor = st.selectbox("Supervisor", opts, format_func=fmt_colaborador)
            sondador = st.selectbox("Sondador", opts, format_func=fmt_colaborador)
            a1, a2 = st.columns(2)
            aux1 = a1.selectbox("Auxiliar 1", opts, format_func=fmt_colaborador)
            aux2 = a2.selectbox("Auxiliar 2", opts, format_func=fmt_colaborador)
            status = st.selectbox("Status", ["Ativa","Inativa"])
            salvar = st.form_submit_button("Cadastrar equipe", type="primary")

        if salvar:
            if codigo.strip() and nome.strip():
                try:
                    execute("""
                        INSERT INTO equipes(
                            codigo,nome,supervisor_id,sondador_id,
                            auxiliar1_id,auxiliar2_id,status
                        )
                        VALUES(?,?,?,?,?,?,?)
                    """, (
                        codigo.strip(), nome.strip(), supervisor,
                        sondador, aux1, aux2, status
                    ))
                    st.success("Equipe cadastrada.")
                    st.rerun()
                except sqlite3.IntegrityError:
                    st.error("Código de equipe já cadastrado.")
            else:
                st.error("Informe código e nome.")

        st.divider()
        st.dataframe(
            query("SELECT * FROM equipes ORDER BY codigo"),
            use_container_width=True,
            hide_index=True
        )

# ============================================================
# SONDAS
# ============================================================
elif page == "🔩 Sondas":
    st.title("🔩 SONDAS")
    dfe = query("SELECT * FROM equipes WHERE status='Ativa' ORDER BY codigo")

    if dfe.empty:
        st.warning("Cadastre uma equipe antes de cadastrar uma sonda.")
    else:
        opts = dfe["id"].tolist()

        def fmt_equipe(x):
            row = dfe[dfe.id == x]
            return str(x) if row.empty else str(row.iloc[0]["codigo"])

        with st.form("form_sonda", clear_on_submit=True):
            c1, c2, c3 = st.columns(3)
            codigo = c1.text_input("Código")
            modelo = c2.text_input("Modelo")
            fabricante = c3.text_input("Fabricante")

            c1, c2, c3 = st.columns(3)
            patrimonio = c1.text_input("Patrimônio")
            equipe = c2.selectbox("Equipe vinculada", opts, format_func=fmt_equipe)
            status = c3.selectbox(
                "Status", ["Operando","Parada","Manutenção","Inativa"]
            )
            salvar = st.form_submit_button("Cadastrar sonda", type="primary")

        if salvar:
            if codigo.strip():
                try:
                    execute("""
                        INSERT INTO sondas(
                            codigo,modelo,fabricante,patrimonio,equipe_id,status
                        )
                        VALUES(?,?,?,?,?,?)
                    """, (
                        codigo.strip(), modelo, fabricante,
                        patrimonio, equipe, status
                    ))
                    st.success("Sonda cadastrada.")
                    st.rerun()
                except sqlite3.IntegrityError:
                    st.error("Código de sonda já cadastrado.")
            else:
                st.error("Informe o código.")

        st.divider()
        st.dataframe(
            query("""
                SELECT s.*,e.codigo equipe
                FROM sondas s
                LEFT JOIN equipes e ON e.id=s.equipe_id
                ORDER BY s.codigo
            """),
            use_container_width=True,
            hide_index=True
        )

# ============================================================
# CADASTROS
# ============================================================
elif page == "⚙️ Cadastros":
    st.title("⚙️ CADASTROS")
    tab_f, tab_a = st.tabs(["🎯 Furos", "⏱️ Códigos de atividades"])

    with tab_f:
        st.subheader("🎯 Cadastro de Furo")

        with st.form("form_furo", clear_on_submit=True):
            c1, c2, c3 = st.columns(3)
            ident = c1.text_input("Identificação do furo")
            projeto = c2.text_input("Projeto")
            cliente = c3.text_input("Cliente")

            c1, c2, c3 = st.columns(3)
            local = c1.text_input("Local")
            e = c2.number_input("Coordenada E")
            n = c3.number_input("Coordenada N")

            c1, c2, c3 = st.columns(3)
            cota = c1.number_input("Cota")
            az = c2.number_input("Azimute")
            dip = c3.number_input("DIP")

            status = st.selectbox(
                "Status", ["Em andamento","Planejado","Concluído"]
            )
            salvar = st.form_submit_button("💾 Cadastrar furo", type="primary")

        if salvar:
            if ident.strip():
                try:
                    execute("""
                        INSERT INTO furos(
                            identificacao,projeto,cliente,local,
                            coord_e,coord_n,cota,azimute,dip,status
                        )
                        VALUES(?,?,?,?,?,?,?,?,?,?)
                    """, (
                        ident.strip(), projeto, cliente, local,
                        e, n, cota, az, dip, status
                    ))
                    st.success("Furo cadastrado com sucesso!")
                    st.rerun()
                except sqlite3.IntegrityError:
                    st.error("Esta identificação de furo já está cadastrada.")
            else:
                st.error("Informe a identificação do furo.")

        st.divider()
        st.subheader("📋 Furos cadastrados")
        df_furos = query("SELECT * FROM furos ORDER BY identificacao")

        if df_furos.empty:
            st.info("Nenhum furo cadastrado.")
        else:
            st.dataframe(df_furos, use_container_width=True, hide_index=True)
            st.divider()
            st.subheader("🗑️ Excluir Furo")

            furo_id_excluir = st.selectbox(
                "Selecione o furo que deseja excluir",
                df_furos["id"].tolist(),
                format_func=lambda x: (
                    str(df_furos[df_furos["id"] == x].iloc[0]["identificacao"])
                )
            )

            furo_sel = df_furos[df_furos["id"] == furo_id_excluir].iloc[0]
            confirmar = st.checkbox(
                f"Confirmo que desejo excluir {furo_sel['identificacao']}"
            )

            if st.button("🗑️ EXCLUIR FURO", use_container_width=True):
                if not confirmar:
                    st.error("Marque a confirmação antes de excluir.")
                else:
                    total = int(query("""
                        SELECT COUNT(*) AS total
                        FROM boletins
                        WHERE furo_id=?
                    """, (int(furo_id_excluir),)).iloc[0]["total"])

                    if total > 0:
                        st.error(
                            f"Não é possível excluir. Existem {total} boletim(ns) vinculados."
                        )
                    else:
                        delete("furos", furo_id_excluir)
                        st.success("Furo excluído com sucesso.")
                        st.rerun()

    with tab_a:
        st.caption("Cadastre ou atualize os códigos de atividades.")
        df = query("SELECT * FROM atividades ORDER BY codigo")
        st.dataframe(df, use_container_width=True, hide_index=True)

        st.divider()
        st.subheader("➕ Cadastrar / Atualizar Atividade")

        with st.form("form_atividade", clear_on_submit=True):
            c1, c2, c3, c4 = st.columns(4)
            cod = c1.number_input("Código", min_value=1, step=1)
            grupo = c2.text_input("Grupo")
            atividade = c3.text_input("Atividade")
            classificacao = c4.selectbox(
                "Classificação",
                [
                    "OPERAÇÃO DIRETA","APOIO OPERACIONAL",
                    "MANUTENÇÃO PREVENTIVA","MECÂNICA CORRETIVA",
                    "PARADA EXTERNA","INTERVENÇÃO NO FURO",
                    "SEGURANÇA","ADMINISTRATIVO"
                ]
            )
            salvar_act = st.form_submit_button(
                "💾 Salvar Atividade", type="primary"
            )

        if salvar_act:
            if grupo.strip() and atividade.strip():
                execute("""
                    INSERT INTO atividades(
                        codigo,grupo,atividade,classificacao
                    )
                    VALUES(?,?,?,?)
                    ON CONFLICT(codigo)
                    DO UPDATE SET
                        grupo=excluded.grupo,
                        atividade=excluded.atividade,
                        classificacao=excluded.classificacao
                """, (
                    int(cod), grupo.strip(),
                    atividade.strip(), classificacao
                ))
                st.success("Atividade cadastrada/atualizada com sucesso.")
                st.rerun()
            else:
                st.error("Preencha todos os campos da atividade.")

# ============================================================
# CADASTRO DE FUROS - PERFIL CAMPO
# ============================================================
elif page == "🕳️ Cadastro de Furos":
    st.title("🕳️ CADASTRO DE FUROS")
    st.caption("Cadastre um novo furo para disponibilizá-lo no preenchimento dos boletins.")

    with st.form("form_furo_campo", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        ident = c1.text_input("Identificação do furo")
        projeto = c2.text_input("Projeto")
        cliente = c3.text_input("Cliente")

        c1, c2, c3 = st.columns(3)
        local = c1.text_input("Local")
        e = c2.number_input("Coordenada E")
        n = c3.number_input("Coordenada N")

        c1, c2, c3 = st.columns(3)
        cota = c1.number_input("Cota")
        az = c2.number_input("Azimute")
        dip = c3.number_input("DIP")

        status = st.selectbox(
            "Status", ["Em andamento", "Planejado", "Concluído"]
        )
        salvar = st.form_submit_button("💾 Cadastrar furo", type="primary")

    if salvar:
        if ident.strip():
            try:
                execute("""
                    INSERT INTO furos(
                        identificacao,projeto,cliente,local,
                        coord_e,coord_n,cota,azimute,dip,status
                    )
                    VALUES(?,?,?,?,?,?,?,?,?,?)
                """, (
                    ident.strip(), projeto, cliente, local,
                    e, n, cota, az, dip, status
                ))
                st.success("Furo cadastrado com sucesso! Ele já pode ser selecionado no Novo Boletim.")
                st.rerun()
            except sqlite3.IntegrityError:
                st.error("Esta identificação de furo já está cadastrada.")
        else:
            st.error("Informe a identificação do furo.")

    st.divider()
    st.subheader("📋 Furos cadastrados")
    df_furos = query("SELECT * FROM furos ORDER BY identificacao")
    if df_furos.empty:
        st.info("Nenhum furo cadastrado.")
    else:
        st.dataframe(df_furos, use_container_width=True, hide_index=True)

# ============================================================
# CONSULTA DE ATIVIDADES - PERFIL CAMPO
# ============================================================
elif page == "📖 Consultar Atividades":
    st.title("📖 CONSULTAR CÓDIGOS DE ATIVIDADES")
    st.caption("Consulta somente. Use esta tela para conferir o código, a atividade e a classificação antes de fazer um lançamento no boletim.")

    busca = st.text_input("🔎 Pesquisar por código, grupo, atividade ou classificação")
    df_atividades = query("SELECT codigo, grupo, atividade, classificacao FROM atividades ORDER BY codigo")

    if df_atividades.empty:
        st.info("Nenhum código de atividade cadastrado.")
    else:
        if busca.strip():
            termo = busca.strip().lower()
            mascara = (
                df_atividades["codigo"].astype(str).str.lower().str.contains(termo, na=False)
                | df_atividades["grupo"].fillna("").astype(str).str.lower().str.contains(termo, na=False)
                | df_atividades["atividade"].fillna("").astype(str).str.lower().str.contains(termo, na=False)
                | df_atividades["classificacao"].fillna("").astype(str).str.lower().str.contains(termo, na=False)
            )
            df_atividades = df_atividades[mascara]

        st.dataframe(df_atividades, use_container_width=True, hide_index=True)
        st.caption(f"{len(df_atividades)} atividade(s) encontrada(s).")

# ============================================================
# GERENCIAMENTO
# ============================================================
elif page == "🛠️ Gerenciamento":
    st.title("🛠️ GERENCIAMENTO OPERACIONAL")
    st.caption("Acompanhe equipes, colaboradores, sondas, produção e boletins.")

    df_equipes = query("""
        SELECT e.*,
               sup.nome AS supervisor_nome,
               son.nome AS sondador_nome,
               aux1.nome AS auxiliar1_nome,
               aux2.nome AS auxiliar2_nome
        FROM equipes e
        LEFT JOIN colaboradores sup ON sup.id=e.supervisor_id
        LEFT JOIN colaboradores son ON son.id=e.sondador_id
        LEFT JOIN colaboradores aux1 ON aux1.id=e.auxiliar1_id
        LEFT JOIN colaboradores aux2 ON aux2.id=e.auxiliar2_id
        ORDER BY e.codigo
    """)

    total_equipes = len(df_equipes)
    equipes_ativas = len(df_equipes[df_equipes["status"] == "Ativa"]) if not df_equipes.empty else 0
    total_sondas = int(query("SELECT COUNT(*) total FROM sondas").iloc[0]["total"])
    total_boletins = int(query("SELECT COUNT(*) total FROM boletins").iloc[0]["total"])

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("👥 EQUIPES", total_equipes)
    c2.metric("🟢 EQUIPES ATIVAS", equipes_ativas)
    c3.metric("🔩 SONDAS", total_sondas)
    c4.metric("📋 BOLETINS", total_boletins)

    tabs = st.tabs([
        "📊 Resumo","👥 Gerenciar Equipes","🔩 Sondas",
        "⛏️ Produção","📋 Boletins"
    ])
    tab_resumo, tab_equipes, tab_sondas, tab_producao, tab_boletins = tabs

    with tab_resumo:
        if df_equipes.empty:
            st.info("Nenhuma equipe cadastrada.")
        else:
            resumo = df_equipes[
                [
                    "codigo","nome","supervisor_nome","sondador_nome",
                    "auxiliar1_nome","auxiliar2_nome","status"
                ]
            ].copy()
            resumo.columns = [
                "Código","Equipe","Supervisor","Sondador",
                "Auxiliar 1","Auxiliar 2","Status"
            ]
            st.dataframe(resumo, use_container_width=True, hide_index=True)

        st.divider()
        sondas_resumo = query("""
            SELECT s.codigo AS sonda,s.modelo,s.fabricante,s.patrimonio,s.status,
                   e.codigo AS equipe,e.nome AS nome_equipe
            FROM sondas s
            LEFT JOIN equipes e ON e.id=s.equipe_id
            ORDER BY e.codigo,s.codigo
        """)
        st.dataframe(sondas_resumo, use_container_width=True, hide_index=True)

    with tab_equipes:
        if df_equipes.empty:
            st.info("Nenhuma equipe cadastrada.")
        else:
            equipe_id = st.selectbox(
                "Selecione a equipe",
                df_equipes["id"].tolist(),
                format_func=lambda x: (
                    f"{df_equipes[df_equipes.id==x].iloc[0]['codigo']} - "
                    f"{df_equipes[df_equipes.id==x].iloc[0]['nome']}"
                ),
                key="gerenciamento_equipe_select"
            )
            equipe = df_equipes[df_equipes["id"] == equipe_id].iloc[0]

            st.subheader(f"👥 {equipe['codigo']} - {equipe['nome']}")
            bcount = int(query(
                "SELECT COUNT(*) total FROM boletins WHERE equipe_id=?",
                (int(equipe_id),)
            ).iloc[0]["total"])
            scount = int(query(
                "SELECT COUNT(*) total FROM sondas WHERE equipe_id=?",
                (int(equipe_id),)
            ).iloc[0]["total"])

            c1, c2, c3 = st.columns(3)
            c1.metric("Status", equipe["status"])
            c2.metric("Boletins", bcount)
            c3.metric("Sondas", scount)

            membros = []
            for campo, funcao, nome_campo in [
                ("supervisor_id","Supervisor","supervisor_nome"),
                ("sondador_id","Sondador","sondador_nome"),
                ("auxiliar1_id","Auxiliar 1","auxiliar1_nome"),
                ("auxiliar2_id","Auxiliar 2","auxiliar2_nome"),
            ]:
                if pd.notna(equipe[campo]):
                    membros.append({
                        "Função": funcao,
                        "Nome": equipe[nome_campo]
                    })

            if membros:
                st.dataframe(pd.DataFrame(membros), use_container_width=True, hide_index=True)

            novo_status = st.selectbox(
                "Novo status",
                ["Ativa","Inativa"],
                index=0 if equipe["status"] == "Ativa" else 1,
                key="gerenciamento_status_equipe"
            )
            if st.button("💾 Atualizar Status da Equipe", type="primary"):
                execute(
                    "UPDATE equipes SET status=? WHERE id=?",
                    (novo_status, int(equipe_id))
                )
                st.rerun()

            confirmar = st.checkbox(
                f"Confirmo que desejo excluir a equipe {equipe['codigo']}"
            )
            if st.button("🗑️ EXCLUIR EQUIPE"):
                if not confirmar:
                    st.error("Confirme a exclusão antes de continuar.")
                elif bcount > 0:
                    st.error(
                        f"Não é possível excluir. Existem {bcount} boletim(ns) vinculados."
                    )
                else:
                    execute(
                        "UPDATE sondas SET equipe_id=NULL WHERE equipe_id=?",
                        (int(equipe_id),)
                    )
                    delete("equipes", equipe_id)
                    st.success("Equipe excluída.")
                    st.rerun()

    with tab_sondas:
        df_sondas = query("""
            SELECT s.id,s.codigo,s.modelo,s.fabricante,s.patrimonio,s.status,
                   e.codigo equipe,e.nome nome_equipe
            FROM sondas s
            LEFT JOIN equipes e ON e.id=s.equipe_id
            ORDER BY s.codigo
        """)

        if df_sondas.empty:
            st.info("Nenhuma sonda cadastrada.")
        else:
            st.dataframe(df_sondas, use_container_width=True, hide_index=True)

            sonda_id = st.selectbox(
                "Selecione uma sonda",
                df_sondas["id"].tolist(),
                format_func=lambda x: str(
                    df_sondas[df_sondas.id==x].iloc[0]["codigo"]
                ),
                key="gerenciamento_sonda_select"
            )
            sonda = df_sondas[df_sondas["id"] == sonda_id].iloc[0]

            novo_status = st.selectbox(
                "Status da sonda",
                ["Operando","Parada","Manutenção","Inativa"],
                index=[
                    "Operando","Parada","Manutenção","Inativa"
                ].index(sonda["status"])
                if sonda["status"] in [
                    "Operando","Parada","Manutenção","Inativa"
                ] else 0,
                key="gerenciamento_status_sonda"
            )

            if st.button("💾 Atualizar Status da Sonda", type="primary"):
                execute(
                    "UPDATE sondas SET status=? WHERE id=?",
                    (novo_status, int(sonda_id))
                )
                st.rerun()

    with tab_producao:
        producao = query("""
            SELECT e.id,e.codigo AS equipe,e.nome,
                   COALESCE(m.metros,0) metros,
                   COALESCE(m.recuperado,0) recuperado,
                   COALESCE(h.horas_operacao,0) horas_operacao
            FROM equipes e
            LEFT JOIN (
                SELECT b.equipe_id,
                       SUM(COALESCE(m.ate_m,0)-COALESCE(m.de_m,0)) metros,
                       SUM(COALESCE(m.recuperado_m,0)) recuperado
                FROM boletins b
                LEFT JOIN manobras m ON m.boletim_id=b.id
                GROUP BY b.equipe_id
            ) m ON m.equipe_id=e.id
            LEFT JOIN (
                SELECT b.equipe_id,
                       SUM(CASE WHEN (TRIM(UPPER(COALESCE(a.classificacao,'')))='OPERAÇÃO DIRETA' OR p.codigo_atividade IN (14,15,16,17,18))
                           THEN COALESCE(p.horas,0) ELSE 0 END) horas_operacao
                FROM boletins b
                LEFT JOIN apontamentos p ON p.boletim_id=b.id
                LEFT JOIN atividades a ON a.codigo=p.codigo_atividade
                GROUP BY b.equipe_id
            ) h ON h.equipe_id=e.id
            ORDER BY metros DESC
        """)

        if producao.empty:
            st.info("Nenhuma produção registrada.")
        else:
            producao["Recuperação %"] = (
                producao["recuperado"] /
                producao["metros"].replace(0, pd.NA) * 100
            ).fillna(0)
            producao["ROP (m/h)"] = (
                producao["metros"] /
                producao["horas_operacao"].replace(0, pd.NA)
            ).fillna(0)

            mostrar = producao[
                [
                    "equipe","nome","metros","recuperado",
                    "Recuperação %","horas_operacao","ROP (m/h)"
                ]
            ].copy()
            mostrar.columns = [
                "Equipe","Nome","Metros","Recuperado",
                "Recuperação %","Horas Operação","ROP (m/h)"
            ]

            st.dataframe(mostrar.round(2), use_container_width=True, hide_index=True)
            st.bar_chart(mostrar[["Equipe","Metros"]].set_index("Equipe"))
            st.bar_chart(
                mostrar[["Equipe","ROP (m/h)"]].set_index("Equipe")
            )

    with tab_boletins:
        if df_equipes.empty:
            st.info("Nenhuma equipe cadastrada.")
        else:
            equipe_boletim_id = st.selectbox(
                "Selecione a equipe",
                df_equipes["id"].tolist(),
                format_func=lambda x: (
                    f"{df_equipes[df_equipes.id==x].iloc[0]['codigo']} - "
                    f"{df_equipes[df_equipes.id==x].iloc[0]['nome']}"
                ),
                key="boletins_por_equipe_select"
            )

            boletins = query("""
                SELECT b.id,b.data,b.turno,b.projeto,b.cliente,
                       s.codigo AS sonda,f.identificacao AS furo,
                       b.horimetro_inicial,b.horimetro_final,b.observacoes
                FROM boletins b
                LEFT JOIN sondas s ON s.id=b.sonda_id
                LEFT JOIN furos f ON f.id=b.furo_id
                WHERE b.equipe_id=?
                ORDER BY b.data DESC,b.id DESC
            """, (int(equipe_boletim_id),))

            if boletins.empty:
                st.info("Nenhum boletim encontrado para esta equipe.")
            else:
                st.dataframe(boletins, use_container_width=True, hide_index=True)
                boletim_id = st.selectbox(
                    "Abrir boletim",
                    boletins["id"].tolist(),
                    format_func=lambda x: f"Boletim #{x}",
                    key="gerenciamento_abrir_boletim_select"
                )
                if st.button("📝 Abrir Boletim", type="primary"):
                    abrir_boletim(int(boletim_id))
                    st.rerun()

# ============================================================
# BACKUP
# ============================================================
elif page == "💾 Backup":
    st.title("💾 BACKUP E SEGURANÇA DO BANCO")
    st.caption("Faça cópias do banco DDH e mantenha uma cópia segura fora do servidor.")

    if nivel != "Administrador":
        st.error("Apenas administradores podem acessar o backup do banco.")
        st.stop()

    if Path(DB).exists():
        tamanho = Path(DB).stat().st_size / 1024
        st.metric("Banco atual", f"{tamanho:.1f} KB")
    else:
        st.warning("O banco ddh.db ainda não foi encontrado.")

    c1, c2 = st.columns(2)
    with c1:
        if st.button("💾 Criar backup agora", type="primary", use_container_width=True):
            arq = criar_backup("manual")
            if arq:
                st.success(f"Backup criado: {arq.name}")
            else:
                st.error("Não foi possível criar o backup porque o banco não foi encontrado.")
    with c2:
        dados = bytes_backup_atual()
        st.download_button(
            "📥 Baixar cópia atual do banco",
            data=dados,
            file_name=f"ddh_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db",
            mime="application/octet-stream",
            use_container_width=True,
            disabled=not bool(dados)
        )

    st.divider()
    st.subheader("📚 Backups disponíveis no servidor")
    backups = sorted(
        [x for x in BACKUP_DIR.glob("*.db")],
        key=lambda x: x.stat().st_mtime,
        reverse=True
    )

    if backups:
        tabela = pd.DataFrame([{
            "Arquivo": x.name,
            "Data": datetime.fromtimestamp(x.stat().st_mtime).strftime("%d/%m/%Y %H:%M:%S"),
            "Tamanho (KB)": round(x.stat().st_size / 1024, 1)
        } for x in backups])
        st.dataframe(tabela, use_container_width=True, hide_index=True)

        escolhido = st.selectbox("Selecionar backup para baixar", [x.name for x in backups])
        arq_escolhido = BACKUP_DIR / escolhido
        st.download_button(
            "📥 Baixar backup selecionado",
            data=arq_escolhido.read_bytes(),
            file_name=arq_escolhido.name,
            mime="application/octet-stream"
        )
    else:
        st.info("Nenhum backup criado ainda.")

    st.divider()
    st.subheader("♻️ Restaurar backup")
    st.warning("A restauração substitui o banco atual. Antes de restaurar, o sistema cria uma cópia de segurança do banco atual.")
    upload = st.file_uploader("Enviar arquivo de backup .db", type=["db"], key="restore_db_upload")
    confirmar = st.checkbox("Confirmo que desejo substituir o banco atual", key="confirmar_restore")
    if st.button("♻️ Restaurar backup enviado", type="secondary", disabled=not(upload and confirmar)):
        with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as tmp:
            tmp.write(upload.getvalue())
            tmp_path = Path(tmp.name)
        try:
            validar_backup_sqlite(tmp_path)
            criar_backup("antes_restauracao")
            os.replace(str(tmp_path), DB)
            st.success("Backup restaurado com sucesso. O aplicativo será reiniciado.")
            st.rerun()
        except Exception as e:
            st.error(f"Não foi possível restaurar o backup: {e}")
        finally:
            if tmp_path.exists():
                try:
                    tmp_path.unlink()
                except OSError:
                    pass

# ============================================================
# USUÁRIOS
# ============================================================
elif page == "👤 Usuários":
    if nivel != "Administrador":
        st.error("Acesso permitido apenas para Administrador.")
        st.stop()

    st.title("👤 USUÁRIOS E PERMISSÕES")
    tab_lista, tab_novo, tab_senha = st.tabs([
        "📋 Usuários","➕ Novo usuário","🔑 Alterar senha"
    ])

    with tab_lista:
        usuarios = query("""
            SELECT u.id,u.nome,u.usuario,u.nivel,u.status,
                   e.codigo AS equipe,u.criado_em
            FROM usuarios u
            LEFT JOIN equipes e ON e.id=u.equipe_id
            ORDER BY u.nome
        """)
        st.dataframe(usuarios, use_container_width=True, hide_index=True)

        if not usuarios.empty:
            uid = st.selectbox(
                "Selecione um usuário para alterar status/excluir",
                usuarios["id"].tolist(),
                format_func=lambda x: (
                    f"{usuarios[usuarios.id==x].iloc[0]['nome']} "
                    f"({usuarios[usuarios.id==x].iloc[0]['usuario']})"
                )
            )
            u = usuarios[usuarios["id"] == uid].iloc[0]

            novo_status = st.selectbox(
                "Status",
                ["Ativo","Inativo"],
                index=0 if u["status"] == "Ativo" else 1
            )

            c1, c2 = st.columns(2)
            if c1.button("💾 Atualizar status", type="primary"):
                execute(
                    "UPDATE usuarios SET status=? WHERE id=?",
                    (novo_status, int(uid))
                )
                st.rerun()

            if c2.button("🗑️ Excluir usuário"):
                if int(uid) == int(st.session_state.usuario_id):
                    st.error("Você não pode excluir o usuário que está logado.")
                elif str(u["usuario"]) == "admin":
                    st.error("O usuário administrador padrão não pode ser excluído.")
                else:
                    delete("usuarios", uid)
                    st.rerun()

    with tab_novo:
        equipes = query("SELECT * FROM equipes WHERE status='Ativa' ORDER BY codigo")
        with st.form("novo_usuario", clear_on_submit=True):
            nome = st.text_input("Nome completo")
            usuario = st.text_input("Usuário")
            senha = st.text_input("Senha inicial", type="password")
            nivel_novo = st.selectbox(
                "Nível",
                ["Administrador","Supervisor","Campo"]
            )

            equipe_id = None
            if not equipes.empty:
                opcoes = [None] + equipes["id"].tolist()
                equipe_id = st.selectbox(
                    "Equipe vinculada (opcional)",
                    opcoes,
                    format_func=lambda x: (
                        "Nenhuma"
                        if x is None
                        else str(equipes[equipes.id==x].iloc[0]["codigo"])
                    )
                )

            salvar_usuario = st.form_submit_button(
                "💾 Criar usuário", type="primary"
            )

        if salvar_usuario:
            if not nome.strip() or not usuario.strip() or not senha:
                st.error("Preencha nome, usuário e senha.")
            else:
                try:
                    execute("""
                        INSERT INTO usuarios(
                            nome,usuario,senha,nivel,equipe_id,status,criado_em
                        )
                        VALUES(?,?,?,?,?,?,?)
                    """, (
                        nome.strip(), usuario.strip(),
                        hash_senha(senha), nivel_novo,
                        equipe_id, "Ativo", datetime.now().isoformat()
                    ))
                    st.success("Usuário criado com sucesso.")
                    st.rerun()
                except sqlite3.IntegrityError:
                    st.error("Este nome de usuário já existe.")

    with tab_senha:
        usuarios = query("SELECT id,nome,usuario FROM usuarios ORDER BY nome")
        uid = st.selectbox(
            "Usuário",
            usuarios["id"].tolist(),
            format_func=lambda x: (
                f"{usuarios[usuarios.id==x].iloc[0]['nome']} "
                f"({usuarios[usuarios.id==x].iloc[0]['usuario']})"
            )
        )
        nova = st.text_input("Nova senha", type="password")
        confirmar = st.text_input("Confirmar nova senha", type="password")

        if st.button("🔑 Salvar nova senha", type="primary"):
            if not nova:
                st.error("Informe a nova senha.")
            elif nova != confirmar:
                st.error("As senhas não conferem.")
            else:
                execute(
                    "UPDATE usuarios SET senha=? WHERE id=?",
                    (hash_senha(nova), int(uid))
                )
                st.success("Senha atualizada com sucesso.")
